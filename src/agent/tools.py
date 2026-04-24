import json
import math
import multiprocessing as mp
import os
from queue import Empty
import re
import tempfile
import time
import traceback
from contextlib import contextmanager
from contextvars import ContextVar
from datetime import datetime, time as datetime_time, timedelta
from difflib import SequenceMatcher
from io import StringIO
from typing import Any, Literal, Optional
from zoneinfo import ZoneInfo

from langchain_core.messages import ToolMessage
from langchain_core.tools import tool
from langgraph.config import get_stream_writer
from langgraph.prebuilt import ToolRuntime
from pydantic import BaseModel, Field

from agent.context_engineering_runtime import CustomContext
from api import supos_kernel_api
from dto import DatabaseConnInfo
from utils import build_chart_document, build_chart_result, build_chart_suite, build_multi_metric_chart_result, logger, normalize_chart_result_item

CURRENT_USERNAME = ContextVar('current_username', default='anonymous')

TOOL_ERROR_CURRENT_CODE_MAX_CHARS = int(os.environ.get('TOOL_ERROR_CURRENT_CODE_MAX_CHARS', '3500'))
TOOL_ERROR_TEXT_MAX_CHARS = int(os.environ.get('TOOL_ERROR_TEXT_MAX_CHARS', '1800'))
TOOL_ERROR_PREVIOUS_ITEM_MAX_CHARS = int(os.environ.get('TOOL_ERROR_PREVIOUS_ITEM_MAX_CHARS', '700'))
TOOL_ERROR_INSTRUCTION_MAX_CHARS = int(os.environ.get('TOOL_ERROR_INSTRUCTION_MAX_CHARS', '700'))
TOOL_ERROR_STDIO_MAX_CHARS = int(os.environ.get('TOOL_ERROR_STDIO_MAX_CHARS', '800'))
TOOL_ERROR_CODE_CONTEXT_MAX_CHARS = int(os.environ.get('TOOL_ERROR_CODE_CONTEXT_MAX_CHARS', '1800'))
GENERATED_CODE_FILENAME = '<generated_analysis_code>'
GENERATED_CODE_CONTEXT_RADIUS = 2

CHART_CONTRACT_ERROR_MARKERS = (
    'structured chart item',
    'chart_document',
    'chart_spec',
    'chart_kind',
    'unsupported chart_kind',
    'charts items',
    'charts item',
    'xaxis',
    'yaxis',
    'series',
)


def _clip_retry_context_text(text: Any, max_chars: int) -> str:
    """Bound retry feedback so repeated tool failures do not exhaust model context."""
    value = str(text or '')
    if max_chars <= 0 or len(value) <= max_chars:
        return value
    suffix = f'\n...[truncated {len(value) - max_chars} chars for retry context budget]'
    return value[:max(0, max_chars - len(suffix))] + suffix


def _clip_retry_context_list(items: list[Any] | None, max_items: int, max_chars: int) -> list[str]:
    return [
        _clip_retry_context_text(item, max_chars)
        for item in (items or [])[-max_items:]
        if str(item or '').strip()
    ]


def _is_chart_contract_error_message(error_message: str) -> bool:
    lowered = (error_message or '').lower()
    if any(marker in lowered for marker in (
        'structured chart item',
        'chart_document',
        'chart_spec',
        'chart_kind',
        'unsupported chart_kind',
        'charts items',
        'charts item',
        'xaxis',
        'yaxis',
    )):
        return True
    return 'series' in lowered and any(marker in lowered for marker in ('chart', 'echarts', 'axis'))


def _build_generated_code_error_context(
    code: str,
    error: Exception,
    traceback_text: str,
) -> dict[str, Any]:
    """Extract the generated-code line that caused a Python runtime error."""
    line_no: int | None = None
    column_no: int | None = None

    if isinstance(error, SyntaxError):
        if isinstance(error.lineno, int):
            line_no = error.lineno
        if isinstance(error.offset, int):
            column_no = error.offset

    traceback_obj = getattr(error, '__traceback__', None)
    if traceback_obj is not None:
        for frame in traceback.extract_tb(traceback_obj):
            if frame.filename in {GENERATED_CODE_FILENAME, '<string>'}:
                line_no = frame.lineno

    code_lines = str(code or '').splitlines()
    context_lines: list[dict[str, Any]] = []
    error_line = ''
    if line_no is not None and 1 <= line_no <= len(code_lines):
        error_line = code_lines[line_no - 1]
        start = max(1, line_no - GENERATED_CODE_CONTEXT_RADIUS)
        end = min(len(code_lines), line_no + GENERATED_CODE_CONTEXT_RADIUS)
        for current_line_no in range(start, end + 1):
            context_lines.append({
                'line_no': current_line_no,
                'code': code_lines[current_line_no - 1],
                'is_error_line': current_line_no == line_no,
            })

    return {
        'filename': GENERATED_CODE_FILENAME,
        'line_no': line_no,
        'column_no': column_no,
        'error_line': error_line,
        'context_lines': context_lines,
        'traceback_excerpt': _clip_retry_context_text(traceback_text, TOOL_ERROR_CODE_CONTEXT_MAX_CHARS),
    }


def _format_error_message_with_code_location(
    error_message: str,
    code_error_context: dict[str, Any] | None,
) -> str:
    line_no = (code_error_context or {}).get('line_no')
    error_line = str((code_error_context or {}).get('error_line') or '').strip()
    if not line_no:
        return error_message

    location = f'错误定位：生成代码第 {line_no} 行'
    if error_line:
        location = f'{location}：{error_line}'
    return f'{error_message}\n{location}'


def _load_data_with_fed_query(sql: str,params: Optional[list[Any]] = None):
    return supos_kernel_api.query_dataframe(sql=sql, params=params)

def _ensure_local_file_exists(file_path: str) -> None:
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"文件不存在: {file_path}")


def load_local_file(file_path: str, sheet_name: Optional[str] = None):
    """
    加载本地 CSV 或 Excel 文件，并返回 pandas DataFrame。

    这个辅助函数属于 `sys_prompt.md` 中约定的 Python 执行工具契约。
    """
    import pandas as pd

    _ensure_local_file_exists(file_path)
    lower_file_path = file_path.lower()

    if lower_file_path.endswith('.csv'):
        return pd.read_csv(file_path)
    if lower_file_path.endswith(('.xlsx', '.xls', '.xlsm')):
        return pd.read_excel(file_path, sheet_name=sheet_name if sheet_name else 0)

    raise ValueError(f"不支持的文件格式: {file_path}")


def _parse_excel_usecols(
    usecols: Optional[list[str] | list[int] | str],
    column_names: list[str],
) -> list[int]:
    if usecols is None:
        return list(range(len(column_names)))

    def _deduplicate(items: list[int]) -> list[int]:
        seen: set[int] = set()
        result: list[int] = []
        for item in items:
            if item not in seen:
                seen.add(item)
                result.append(item)
        return result

    if isinstance(usecols, list):
        if all(isinstance(item, int) for item in usecols):
            resolved = [int(item) for item in usecols if 0 <= int(item) < len(column_names)]
            if not resolved:
                raise ValueError('低内存读取时没有匹配到任何 Excel 列，请检查 usecols 配置。')
            return _deduplicate(resolved)
        requested_names = {str(item).strip() for item in usecols if str(item).strip()}
        resolved = [index for index, name in enumerate(column_names) if name in requested_names]
        if not resolved:
            raise ValueError('低内存读取时没有匹配到任何 Excel 列，请检查 usecols 配置。')
        return _deduplicate(resolved)

    if isinstance(usecols, str):
        normalized = usecols.strip()
        if not normalized:
            return list(range(len(column_names)))

        try:
            from openpyxl.utils.cell import column_index_from_string
        except ImportError:
            column_index_from_string = None

        if column_index_from_string and ':' in normalized:
            start_token, end_token = [part.strip() for part in normalized.split(':', 1)]
            if start_token.isalpha() and end_token.isalpha():
                start_index = column_index_from_string(start_token.upper()) - 1
                end_index = column_index_from_string(end_token.upper()) - 1
                return list(range(max(start_index, 0), min(end_index, len(column_names) - 1) + 1))

        tokens = [token.strip() for token in normalized.split(',') if token.strip()]
        if tokens and all(token.isdigit() for token in tokens):
            resolved = [int(token) for token in tokens if 0 <= int(token) < len(column_names)]
            if not resolved:
                raise ValueError('低内存读取时没有匹配到任何 Excel 列，请检查 usecols 配置。')
            return _deduplicate(resolved)
        if column_index_from_string and tokens and all(token.isalpha() for token in tokens):
            resolved = [column_index_from_string(token.upper()) - 1 for token in tokens]
            resolved = [index for index in resolved if 0 <= index < len(column_names)]
            if not resolved:
                raise ValueError('低内存读取时没有匹配到任何 Excel 列，请检查 usecols 配置。')
            return _deduplicate(resolved)

        requested_names = set(tokens)
        resolved = [index for index, name in enumerate(column_names) if name in requested_names]
        if not resolved:
            raise ValueError('低内存读取时没有匹配到任何 Excel 列，请检查 usecols 配置。')
        return _deduplicate(resolved)

    raise ValueError('低内存读取时暂不支持当前 usecols 类型，请改用列名列表、列序号列表或逗号分隔字符串。')


def _apply_batch_transforms(
    dataframe,
    dtype: Optional[dict[str, Any] | str] = None,
    parse_dates: Optional[list[str] | bool] = None,
):
    import pandas as pd

    if parse_dates is True:
        for column in dataframe.columns:
            dataframe[column] = pd.to_datetime(dataframe[column], errors='ignore')
    elif isinstance(parse_dates, list):
        for column in parse_dates:
            if column in dataframe.columns:
                dataframe[column] = pd.to_datetime(dataframe[column], errors='coerce')

    if dtype is None:
        return dataframe

    if isinstance(dtype, dict):
        matched_dtype = {column: value for column, value in dtype.items() if column in dataframe.columns}
        if matched_dtype:
            return dataframe.astype(matched_dtype)
        return dataframe

    return dataframe.astype(dtype)


def _iter_excel_rows_in_batches(
    file_path: str,
    sheet_name: Optional[str] = None,
    chunk_size: int = 50000,
    usecols: Optional[list[str] | list[int] | str] = None,
    dtype: Optional[dict[str, Any] | str] = None,
    parse_dates: Optional[list[str] | bool] = None,
):
    import pandas as pd

    normalized_chunk_size = max(int(chunk_size or 50000), 1)
    lower_file_path = file_path.lower()

    if lower_file_path.endswith(('.xlsx', '.xlsm')):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ValueError('当前环境缺少 openpyxl，无法对 .xlsx/.xlsm 文件执行低内存分批读取。') from exc

        workbook = load_workbook(file_path, read_only=True, data_only=True)
        try:
            worksheet = workbook[sheet_name] if sheet_name else workbook.worksheets[0]
            row_iterator = worksheet.iter_rows(values_only=True)
            headers = next(row_iterator, None)
            if headers is None:
                return

            column_names = [str(value).strip() if value is not None else '' for value in headers]
            selected_indexes = _parse_excel_usecols(usecols, column_names)
            selected_columns = [column_names[index] or f'column_{index + 1}' for index in selected_indexes]
            batch_rows: list[list[Any]] = []

            for row in row_iterator:
                current_row = row or ()
                batch_rows.append([
                    current_row[index] if index < len(current_row) else None
                    for index in selected_indexes
                ])
                if len(batch_rows) >= normalized_chunk_size:
                    batch_frame = pd.DataFrame(batch_rows, columns=selected_columns)
                    yield _apply_batch_transforms(batch_frame, dtype=dtype, parse_dates=parse_dates)
                    batch_rows = []

            if batch_rows:
                batch_frame = pd.DataFrame(batch_rows, columns=selected_columns)
                yield _apply_batch_transforms(batch_frame, dtype=dtype, parse_dates=parse_dates)
        finally:
            workbook.close()
        return

    if lower_file_path.endswith('.xls'):
        try:
            import xlrd
        except ImportError as exc:
            raise ValueError('当前环境缺少 xlrd，无法对 .xls 文件执行低内存分批读取；必要时请提示用户转换为 .xlsx 或 CSV。') from exc

        workbook = xlrd.open_workbook(file_path, on_demand=True)
        try:
            worksheet = workbook.sheet_by_name(sheet_name) if sheet_name else workbook.sheet_by_index(0)
            if worksheet.nrows == 0:
                return

            column_names = [str(value).strip() if value is not None else '' for value in worksheet.row_values(0)]
            selected_indexes = _parse_excel_usecols(usecols, column_names)
            selected_columns = [column_names[index] or f'column_{index + 1}' for index in selected_indexes]

            for start_row in range(1, worksheet.nrows, normalized_chunk_size):
                end_row = min(start_row + normalized_chunk_size, worksheet.nrows)
                batch_rows = []
                for row_index in range(start_row, end_row):
                    row_values = worksheet.row_values(row_index)
                    batch_rows.append([
                        row_values[index] if index < len(row_values) else None
                        for index in selected_indexes
                    ])
                if batch_rows:
                    batch_frame = pd.DataFrame(batch_rows, columns=selected_columns)
                    yield _apply_batch_transforms(batch_frame, dtype=dtype, parse_dates=parse_dates)
        finally:
            workbook.release_resources()
        return

    raise ValueError(f'不支持的 Excel 文件格式: {file_path}')


def load_local_file_low_memory(
    file_path: str,
    sheet_name: Optional[str] = None,
    chunk_size: int = 50000,
    usecols: Optional[list[str] | list[int] | str] = None,
    dtype: Optional[dict[str, Any] | str] = None,
    parse_dates: Optional[list[str] | bool] = None,
    low_memory: bool = True,
):
    """按批次低内存读取本地 CSV 或 Excel，适合 OOM 后的修复性重试场景。"""
    import pandas as pd

    _ensure_local_file_exists(file_path)
    lower_file_path = file_path.lower()
    normalized_chunk_size = max(int(chunk_size or 50000), 1)

    if lower_file_path.endswith('.csv'):
        return pd.read_csv(
            file_path,
            chunksize=normalized_chunk_size,
            usecols=usecols,
            dtype=dtype,
            parse_dates=parse_dates,
            low_memory=low_memory,
        )

    if lower_file_path.endswith(('.xlsx', '.xls', '.xlsm')):
        return _iter_excel_rows_in_batches(
            file_path=file_path,
            sheet_name=sheet_name,
            chunk_size=normalized_chunk_size,
            usecols=usecols,
            dtype=dtype,
            parse_dates=parse_dates,
        )

    raise ValueError(f"不支持的文件格式: {file_path}")


def load_minio_file(bucket: str, object_name: str, sheet_name: Optional[str] = None):
    """预留给后续 MinIO 文件加载场景的辅助函数。"""
    pass


def load_data_with_sql(sql: str, params: Optional[list[Any]] = None):
   return _load_data_with_fed_query(sql=sql,params=params)

def load_local_data_with_sql(sql: str, params: Optional[list[Any]] = None):
    """通过 SQL 从当前配置的数据源中加载表格数据。"""
    import pandas as pd

    from config.database import engine

    with engine.connect() as connection:
        return pd.read_sql(sql, connection, params=params if params else None)

def load_data_with_api(
    endpoint: str,
    method: str = "GET",
    params: Optional[dict[str, Any]] = None,
    headers: Optional[dict[str, Any]] = None,
    timeout: int = 30,
):
    """通过 HTTP API 加载表格数据，并返回 pandas DataFrame。"""
    from io import StringIO as CsvBuffer

    import pandas as pd
    import requests

    response = requests.request(
        method=method.upper(),
        url=endpoint,
        params=params if method.upper() == "GET" else None,
        json=params if method.upper() == "POST" else None,
        headers=headers,
        timeout=timeout,
    )
    response.raise_for_status()

    content_type = response.headers.get('Content-Type', '')
    if 'application/json' in content_type:
        json_data = response.json()
        if isinstance(json_data, list):
            return pd.DataFrame(json_data)
        if isinstance(json_data, dict):
            if 'data' in json_data and isinstance(json_data['data'], list):
                return pd.DataFrame(json_data['data'])
            return pd.DataFrame([json_data])
        raise ValueError("不支持的 JSON 数据格式")

    if 'text/csv' in content_type or endpoint.endswith('.csv'):
        return pd.read_csv(CsvBuffer(response.text))

    raise ValueError(f"不支持的响应格式: {content_type}")


def get_day_range(days_ago: int = 0, timezone_name: str = 'Asia/Shanghai') -> tuple[datetime, datetime]:
    """
    返回某个自然日的起止时间。

    `days_ago=0` 表示今天，`1` 表示昨天，`2` 表示前天。
    返回值始终带时区信息，便于与带时区的时间列直接比较。
    """
    tz = ZoneInfo(timezone_name)
    target_date = (datetime.now(tz) - timedelta(days=days_ago)).date()
    start_at = datetime.combine(target_date, datetime_time.min, tzinfo=tz)
    end_at = datetime.combine(target_date, datetime_time.max, tzinfo=tz)
    return start_at, end_at


def describe_time_coverage(dataframe, column_name: str) -> dict[str, Any]:
    """
    描述某个时间列的数据覆盖范围。

    适合在“最近 / 上个月 / 近 N 天 / 根因关联”这类时间敏感分析中，
    先判断目标时间窗口是否命中数据，再决定后续过滤、关联和空结果处理策略。
    """
    import pandas as pd

    if dataframe is None:
        return {
            'column_name': column_name,
            'row_count': 0,
            'non_null_count': 0,
            'min_time': None,
            'max_time': None,
        }

    if not isinstance(dataframe, pd.DataFrame):
        dataframe = pd.DataFrame(dataframe)

    if column_name not in dataframe.columns:
        raise KeyError(column_name)

    series = pd.to_datetime(dataframe[column_name], errors='coerce')
    valid_series = series.dropna()
    if valid_series.empty:
        return {
            'column_name': column_name,
            'row_count': int(len(dataframe)),
            'non_null_count': 0,
            'min_time': None,
            'max_time': None,
        }

    return {
        'column_name': column_name,
        'row_count': int(len(dataframe)),
        'non_null_count': int(valid_series.shape[0]),
        'min_time': valid_series.min().isoformat(),
        'max_time': valid_series.max().isoformat(),
    }


def probe_distinct_values(
    dataframe,
    column_name: str,
    top_n: int = 20,
    dropna: bool = True,
) -> list[dict[str, Any]]:
    """
    返回某个字段的高频取值分布，适合在“无数据重试”时先做轻量探测。

    常见用途：
    - 看筛选字段是否真的存在预期取值
    - 看枚举值是否存在空格、大小写、前后缀差异
    - 给后续条件纠偏提供候选值，而不是盲目放宽语义
    """
    import pandas as pd

    frame = _ensure_probe_dataframe(dataframe)
    if column_name not in frame.columns:
        raise KeyError(column_name)

    normalized_top_n = max(int(top_n or 20), 1)
    series = frame[column_name]
    counts = series.value_counts(dropna=dropna).head(normalized_top_n)

    return [
        {
            'value': _serialize_probe_value(index),
            'count': int(count),
            'normalized_value': _normalize_probe_text(index),
        }
        for index, count in counts.items()
    ]


def probe_text_candidates(
    dataframe,
    column_name: str,
    keyword: str,
    top_n: int = 10,
) -> list[dict[str, Any]]:
    """
    基于字符串规范化和相似度，为目标关键词返回候选值。

    适合在无数据重试时，用于判断：
    - 用户输入值是否存在轻微格式差异
    - 是否应该把精确匹配纠偏成 contains / LIKE
    - 是否只是空格、大小写、连接符、中英文括号等表达差异
    """
    import pandas as pd

    frame = _ensure_probe_dataframe(dataframe)
    if column_name not in frame.columns:
        raise KeyError(column_name)

    normalized_keyword = _normalize_probe_text(keyword)
    if not normalized_keyword:
        return []

    counts = frame[column_name].value_counts(dropna=True)
    candidates: list[dict[str, Any]] = []
    for raw_value, count in counts.items():
        normalized_value = _normalize_probe_text(raw_value)
        if not normalized_value:
            continue
        similarity = SequenceMatcher(None, normalized_keyword, normalized_value).ratio()
        exact_match = normalized_value == normalized_keyword
        contains_match = normalized_keyword in normalized_value or normalized_value in normalized_keyword
        if not exact_match and not contains_match and similarity < 0.45:
            continue

        if exact_match:
            match_type = 'exact'
            score = 1.0
        elif contains_match:
            match_type = 'contains'
            score = max(similarity, 0.8)
        else:
            match_type = 'similar'
            score = similarity

        candidates.append({
            'value': _serialize_probe_value(raw_value),
            'count': int(count),
            'normalized_value': normalized_value,
            'match_type': match_type,
            'similarity': round(float(score), 4),
        })

    candidates.sort(
        key=lambda item: (
            {'exact': 3, 'contains': 2, 'similar': 1}.get(item['match_type'], 0),
            item['similarity'],
            item['count'],
        ),
        reverse=True,
    )
    return candidates[:max(int(top_n or 10), 1)]


def _ensure_probe_dataframe(dataframe):
    import pandas as pd

    if dataframe is None:
        return pd.DataFrame()
    if isinstance(dataframe, pd.DataFrame):
        return dataframe
    return pd.DataFrame(dataframe)


def _normalize_probe_text(value: Any) -> str:
    if value is None:
        return ''
    text = str(value).strip().lower()
    if not text or text == 'nan':
        return ''
    text = text.replace('（', '(').replace('）', ')')
    text = text.replace('【', '[').replace('】', ']')
    text = text.replace('－', '-').replace('—', '-').replace('_', '-')
    text = re.sub(r'\s+', '', text)
    text = re.sub(r'[-/]+', '-', text)
    return text


def _serialize_probe_value(value: Any) -> Any:
    try:
        import pandas as pd

        if pd.isna(value):
            return None
    except Exception:
        pass
    return value.item() if hasattr(value, 'item') else value


def build_markdown_table(dataframe, columns: Optional[list[str]] = None, max_rows: int = 10) -> str:
    """
    把 DataFrame 转成 Markdown 表格文本。

    这个辅助函数适合在“需要输出 Markdown 表格”的场景中直接复用，
    用来减少手写表格拼接时的引号、换行和格式错误。
    如果当前分析任务并不需要表格展示，则不必使用它。
    """
    import pandas as pd

    if dataframe is None:
        return ''

    if not isinstance(dataframe, pd.DataFrame):
        dataframe = pd.DataFrame(dataframe)

    if columns:
        available_columns = [column for column in columns if column in dataframe.columns]
        table_df = dataframe[available_columns].copy()
    else:
        table_df = dataframe.copy()

    if max_rows > 0:
        table_df = table_df.head(max_rows)

    if table_df.empty:
        return ''

    table_df = table_df.fillna('')

    def _stringify(value: Any) -> str:
        if isinstance(value, float):
            return f"{value:,.2f}".rstrip('0').rstrip('.')
        return str(value)

    headers = [str(column) for column in table_df.columns]
    header_line = "| " + " | ".join(headers) + " |"
    separator_line = "| " + " | ".join(['---'] * len(headers)) + " |"

    body_lines = []
    for _, row in table_df.iterrows():
        values = [_stringify(row[column]).replace('|', '&#124;') for column in table_df.columns]
        body_lines.append("| " + " | ".join(values) + " |")

    return "\n".join([header_line, separator_line, *body_lines])


class StructuredResult(BaseModel):
    """当前 Agent 运行时消费的标准工具返回结构。"""

    analysis_report: str = Field(description="分析报告内容")
    charts: list[dict[str, Any]] = Field(default_factory=list, description="图表产物列表")
    tables: list[dict[str, Any]] = Field(default_factory=list, description="表格产物列表")


class RetryResult(BaseModel):
    """Structured feedback that asks the agent to generate another code attempt."""

    retry_type: Literal[
        "probe_feedback",
        "data_source_unavailable",
        "missing_structured_artifact",
        "missing_report",
        "incomplete_result",
        "contract_error",
        "chart_contract_error",
    ] = Field(description="Why the current execution needs another attempt.")
    message: str = Field(description="Human-readable retry reason.")
    diagnostics: dict[str, Any] = Field(default_factory=dict, description="Structured diagnostic facts.")
    repair_instructions: list[str] = Field(default_factory=list, description="Targeted next-step instructions.")
    analysis_report: str = Field(default="", description="Optional diagnostic report; not a final answer.")


class NoDataFoundError(Exception):
    """用于把“当前代码筛选后无数据”的情况显式上抛给 execute_python。"""

    def __init__(self, reason: str, detail_lines: Optional[list[str]] = None):
        normalized_reason = (reason or '当前条件下未命中可分析数据。').strip() or '当前条件下未命中可分析数据。'
        normalized_details = [
            str(item).strip()
            for item in (detail_lines or [])
            if str(item).strip()
        ]
        self.reason = normalized_reason
        self.detail_lines = normalized_details
        super().__init__(normalized_reason)


def raise_no_data_error(reason: str, detail_lines: Optional[list[str]] = None) -> None:
    """
    在数据加载、过滤、关联或聚合后发现结果为空时，显式通知上层重试。

    生成代码遇到这类情况时，不应继续生成空图表并调用 save_analysis_result() 收口，
    而应调用本函数把“本轮未命中数据”的信息返回给 execute_python，再由上层决定是否继续重试。
    """
    raise NoDataFoundError(reason=reason, detail_lines=detail_lines)


def request_retry(
    retry_type: str,
    message: str,
    diagnostics: Optional[dict[str, Any]] = None,
    repair_instructions: Optional[list[str]] = None,
    analysis_report: str = "",
) -> RetryResult:
    """Return structured retry feedback from generated probe code."""
    allowed_types = {
        "probe_feedback",
        "data_source_unavailable",
        "missing_structured_artifact",
        "missing_report",
        "incomplete_result",
        "contract_error",
        "chart_contract_error",
    }
    normalized_retry_type = str(retry_type or "").strip() or "contract_error"
    if normalized_retry_type not in allowed_types:
        normalized_retry_type = "contract_error"

    normalized_message = str(message or "").strip() or "当前代码执行完成，但仍需要重新生成代码继续分析。"
    normalized_instructions = [
        str(item).strip()
        for item in (repair_instructions or [])
        if str(item).strip()
    ]
    return RetryResult(
        retry_type=normalized_retry_type,  # type: ignore[arg-type]
        message=normalized_message,
        diagnostics=diagnostics or {},
        repair_instructions=normalized_instructions,
        analysis_report=str(analysis_report or "").strip(),
    )


def save_empty_analysis_result(
    title: str,
    reason: str,
    detail_lines: Optional[list[str]] = None,
) -> StructuredResult:
    """
    在目标时间范围或关联结果为空时，生成一份可直接展示的空结果分析。

    这个函数保留给兼容场景使用。

    对于当前主链路，如果只是本轮 SQL、时间窗、筛选条件或关联条件没有命中数据，
    应优先调用 raise_no_data_error(...) 把“无数据”返回给上层重试，而不是直接在 Python 层收口。
    """
    import html

    safe_title = (title or '分析结果').strip() or '分析结果'
    safe_reason = (reason or '当前条件下未命中可分析数据。').strip() or '当前条件下未命中可分析数据。'
    lines = [str(item).strip() for item in (detail_lines or []) if str(item).strip()]

    report_sections = [
        f"## {safe_title}",
        "### 结果说明",
        f"- {safe_reason}",
    ]
    if lines:
        report_sections.extend([
            "### 补充信息",
            *[f"- {item}" for item in lines],
        ])

    analysis_report = '\n\n'.join(section for section in report_sections if section)
    return save_analysis_result(
        analysis_report=analysis_report,
        charts=[],
        tables=[],
    )


def _normalize_chart_items(
    charts: Optional[list[dict[str, Any]]] = None,
) -> list[dict[str, Any]]:
    """统一图表结果结构。"""
    normalized_items: list[dict[str, Any]] = []

    for index, chart in enumerate(charts or [], start=1):
        normalized_items.append(normalize_chart_result_item(chart, index=index))

    return normalized_items


def _normalize_table_items(tables: Optional[list[dict[str, Any]]] = None) -> list[dict[str, Any]]:
    """统一表格结果结构。"""
    normalized_items: list[dict[str, Any]] = []
    for index, table in enumerate(tables or [], start=1):
        if not isinstance(table, dict):
            raise ValueError('tables 中的每一项都必须是 dict。')

        title = str(table.get('title') or f'表格 {index}').strip()
        description = str(table.get('description') or '').strip()
        columns = table.get('columns') or []
        rows = table.get('rows') or []
        if not isinstance(columns, list) or not isinstance(rows, list):
            raise ValueError('表格产物必须包含 columns(list) 与 rows(list)。')

        normalized_items.append({
            "title": title,
            "description": description,
            "columns": columns,
            "rows": rows,
        })
    return normalized_items


def save_analysis_result(
    analysis_report: str = '',
    charts: Optional[list[dict[str, Any]]] = None,
    tables: Optional[list[dict[str, Any]]] = None,
) -> StructuredResult:
    """
    结束一次由生成代码驱动的分析任务，并返回标准结构化结果。

    当前主契约是：
    - analysis_report: Markdown 报告
    - charts: 图表数组
    - tables: 表格数组

    """
    report_text = (analysis_report or '').strip()
    if not report_text:
        raise ValueError('analysis_report 不能为空，必须传入完整的 Markdown 分析报告。')

    normalized_charts = _normalize_chart_items(charts=charts)
    normalized_tables = _normalize_table_items(tables=tables)
    return StructuredResult(
        analysis_report=report_text,
        charts=normalized_charts,
        tables=normalized_tables,
    )


class ExePythonCodeInput(BaseModel):
    code: str = Field(description="待执行的 Python 代码")
    title: str = Field(description="本次代码执行任务的简短标题")
    description: str = Field(description="本次代码执行任务的详细说明")


def _serialize_exec_result(exec_result: Any) -> dict[str, Any]:
    if isinstance(exec_result, StructuredResult):
        return {
            'result_kind': 'structured',
            'payload': exec_result.model_dump(),
        }
    if isinstance(exec_result, RetryResult):
        return {
            'result_kind': 'retry',
            'payload': exec_result.model_dump(),
        }
    return {
        'result_kind': 'raw',
        'payload': None,
    }


def _deserialize_exec_result(payload: dict[str, Any]) -> Any:
    if payload.get('result_kind') == 'structured' and isinstance(payload.get('payload'), dict):
        return StructuredResult(**payload['payload'])
    if payload.get('result_kind') == 'retry' and isinstance(payload.get('payload'), dict):
        return RetryResult(**payload['payload'])
    return None


def _build_no_data_feedback(
    reason: str,
    detail_lines: Optional[list[str]] = None,
) -> dict[str, Any]:
    normalized_reason = (reason or '当前条件下未命中可分析数据。').strip() or '当前条件下未命中可分析数据。'
    normalized_details = [
        str(item).strip()
        for item in (detail_lines or [])
        if str(item).strip()
    ]
    error_message = normalized_reason
    if normalized_details:
        error_message = f"{normalized_reason} 细节：{'；'.join(normalized_details[:5])}"
    return {
        'error_type': 'no_data_found',
        'error_message': error_message,
        'error_signature': {
            'error_type': 'no_data_found',
            'signature_type': 'no_data_found',
            'reason': normalized_reason,
            'detail_lines': normalized_details[:5],
        },
        'repair_instructions': _build_repair_instructions('no_data_found'),
    }


def _build_retry_error_signature(retry_result: RetryResult) -> dict[str, Any]:
    return {
        'error_type': retry_result.retry_type,
        'signature_type': retry_result.retry_type,
        'message': retry_result.message,
        'diagnostics': retry_result.diagnostics,
    }


def _is_invalid_chart_data_value(value: Any) -> bool:
    """Return True when a chart data point cannot be rendered by ECharts."""
    if value is None:
        return True
    if isinstance(value, float):
        return math.isnan(value) or math.isinf(value)
    if hasattr(value, "item"):
        try:
            return _is_invalid_chart_data_value(value.item())
        except Exception:
            return False
    if isinstance(value, dict):
        return any(_is_invalid_chart_data_value(item) for item in value.values())
    if isinstance(value, (list, tuple)):
        return any(_is_invalid_chart_data_value(item) for item in value)
    return False


def _find_invalid_chart_data_points(charts: list[dict[str, Any]] | None) -> list[dict[str, Any]]:
    invalid_points: list[dict[str, Any]] = []
    for chart_index, chart in enumerate(charts or []):
        if not isinstance(chart, dict):
            continue
        chart_spec = chart.get("chart_spec") or {}
        if not isinstance(chart_spec, dict):
            continue
        series_items = chart_spec.get("series") or []
        if isinstance(series_items, dict):
            series_items = [series_items]
        if not isinstance(series_items, list):
            continue
        for series_index, series in enumerate(series_items):
            if not isinstance(series, dict):
                continue
            data_items = series.get("data") or []
            if not isinstance(data_items, list):
                continue
            for data_index, value in enumerate(data_items):
                if _is_invalid_chart_data_value(value):
                    invalid_points.append({
                        "chart_index": chart_index,
                        "chart_title": str(chart.get("title") or ""),
                        "series_index": series_index,
                        "series_name": str(series.get("name") or ""),
                        "data_index": data_index,
                        "value_repr": repr(value),
                    })
                    if len(invalid_points) >= 10:
                        return invalid_points
    return invalid_points


def _validate_structured_result_contract(exec_result: StructuredResult) -> RetryResult | None:
    report_text = (exec_result.analysis_report or '').strip()
    if not report_text:
        return request_retry(
            retry_type="missing_report",
            message="当前代码返回了结构化结果，但 analysis_report 为空。",
            diagnostics={
                "chart_count": len(exec_result.charts or []),
                "table_count": len(exec_result.tables or []),
            },
            repair_instructions=[
                "下一版代码必须先构造完整 Markdown 分析报告。",
                "调用 save_analysis_result(...) 时传入非空 analysis_report。",
            ],
        )

    if not (exec_result.charts or exec_result.tables):
        return request_retry(
            retry_type="missing_structured_artifact",
            message="当前代码只生成了分析报告，没有生成图表或结构化表格。",
            diagnostics={
                "analysis_report_preview": report_text[:800],
                "chart_count": 0,
                "table_count": 0,
            },
            repair_instructions=[
                "下一版代码保留已有统计口径、筛选条件和核心结论。",
                "至少补充一个 charts 或 tables 结构化产物。",
                "趋势、对比、TopN 等场景优先生成 charts；单指标或明细汇总可以生成 tables。",
                "不要把自然语言报告直接作为最终结果结束。",
            ],
            analysis_report=report_text,
        )

    invalid_chart_points = _find_invalid_chart_data_points(exec_result.charts)
    if invalid_chart_points:
        return request_retry(
            retry_type="chart_contract_error",
            message="当前代码生成了图表，但图表数据中包含 None/null/NaN/Inf 等不可展示值，可能导致前端图表空白。",
            diagnostics={
                "analysis_report_preview": report_text[:800],
                "chart_count": len(exec_result.charts or []),
                "table_count": len(exec_result.tables or []),
                "invalid_chart_points": invalid_chart_points,
            },
            repair_instructions=[
                "下一版代码保留已有统计口径、筛选条件和核心结论。",
                "把所有传入 charts/chart_spec/series.data 的统计值转换为 Python 原生 int、float 或 str。",
                "不要把 numpy.int64、numpy.float64、pandas 标量、NaN、Inf 或 None 直接传给 pyecharts/ECharts。",
                "典型修复：total_sales = int(total_sales)，或 add_yaxis(..., [int(total_sales)])。",
                "生成图表前检查 series.data，确保不存在 None/null/NaN/Inf。",
            ],
            analysis_report=report_text,
        )

    return None


def _classify_process_exit(process_exitcode: int | None) -> tuple[str, str]:
    normalized_exitcode = int(process_exitcode) if process_exitcode is not None else 0
    if normalized_exitcode in (-9, 137):
        return (
            'resource_exhausted',
            f'代码执行子进程异常退出，exitcode={normalized_exitcode}。这通常意味着内存耗尽或被系统因资源压力终止，请按大文件/低内存策略重写加载逻辑。',
        )
    return (
        'runtime_error',
        f'代码执行子进程异常退出，exitcode={normalized_exitcode}',
    )


def _build_execution_result_temp_dir() -> str:
    """返回执行结果临时目录，并确保目录存在。"""
    from config.config import Config

    base_dir = getattr(Config, 'TEMP_DIR', '') or tempfile.gettempdir()
    result_dir = os.path.join(base_dir, 'python_exec_results')
    os.makedirs(result_dir, exist_ok=True)
    return result_dir


def _write_worker_result_payload(payload: dict[str, Any]) -> str:
    """将完整执行结果落到临时文件，避免大对象阻塞 multiprocessing.Queue。"""
    result_dir = _build_execution_result_temp_dir()
    with tempfile.NamedTemporaryFile(
        mode='w',
        encoding='utf-8',
        suffix='.json',
        prefix='python_exec_',
        dir=result_dir,
        delete=False,
    ) as temp_file:
        json.dump(payload, temp_file, ensure_ascii=False)
        temp_file.flush()
        return temp_file.name


def _read_worker_result_payload(result_file_path: str) -> dict[str, Any]:
    with open(result_file_path, 'r', encoding='utf-8') as result_file:
        return json.load(result_file)


def _remove_worker_result_payload(result_file_path: str) -> None:
    if result_file_path:
        try:
            os.remove(result_file_path)
        except FileNotFoundError:
            pass


def _sanitize_tool_output(output: str) -> str:
    """从 stdout/stderr 中去掉流式事件碎片，只保留有意义的执行日志。"""
    if not output:
        return ''

    cleaned_lines: list[str] = []
    for line in output.splitlines():
        stripped = line.strip()
        if stripped.startswith("data: {") and '"type"' in stripped:
            continue
        if stripped.startswith("{'type':") and "'stage':" in stripped:
            continue
        cleaned_lines.append(line)
    return "\n".join(cleaned_lines).strip()


def _create_execution_record(
    conversation_id: int,
    turn_id: int,
    tool_call_id: str,
    code: str,
    title: str,
    description: str,
) -> int:
    """在生成代码真正执行前，先创建一条执行记录。"""
    from config.database import SessionLocal
    from model import InsightNsExecution

    session = SessionLocal()
    try:
        execution = InsightNsExecution(
            conversation_id=conversation_id,
            turn_id=turn_id,
            tool_call_id=tool_call_id or '',
            title=(title or '')[:255],
            description=description or '',
            generated_code=code or '',
            execution_status='running',
        )
        session.add(execution)
        session.commit()
        session.refresh(execution)
        return int(execution.id)
    finally:
        session.close()


def _update_execution_record(
    execution_id: int,
    *,
    execution_status: str,
    analysis_report: str = '',
    result_payload_json: str = '',
    stdout_text: str = '',
    stderr_text: str = '',
    execution_seconds: int = 0,
    error_message: str = '',
) -> None:
    """在成功、失败或结果不合法后，回写执行记录。"""
    if execution_id <= 0:
        return

    from config.database import SessionLocal
    from model import InsightNsExecution

    session = SessionLocal()
    try:
        execution = session.query(InsightNsExecution).filter(
            InsightNsExecution.id == execution_id,
            InsightNsExecution.is_deleted == 0,
        ).first()
        if execution is None:
            return

        execution.execution_status = execution_status
        execution.analysis_report = analysis_report or ''
        execution.result_payload_json = result_payload_json or '{}'
        execution.stdout_text = stdout_text or ''
        execution.stderr_text = stderr_text or ''
        execution.execution_seconds = int(execution_seconds or 0)
        execution.error_message = error_message or ''
        execution.updated_at = datetime.now()
        execution.finished_at = datetime.now()
        session.commit()
    finally:
        session.close()


def _build_execution_namespace() -> dict[str, Any]:
    """向生成代码暴露允许使用的辅助函数命名空间。"""
    return {
        'load_local_file': load_local_file,
        'load_local_file_low_memory': load_local_file_low_memory,
        'load_minio_file': load_minio_file,
        'load_data_with_sql': load_data_with_sql,
        'load_data_with_api': load_data_with_api,
        'get_day_range': get_day_range,
        'describe_time_coverage': describe_time_coverage,
        'probe_distinct_values': probe_distinct_values,
        'probe_text_candidates': probe_text_candidates,
        'build_markdown_table': build_markdown_table,
        'build_chart_document': build_chart_document,
        'build_chart_result': build_chart_result,
        'build_chart_suite': build_chart_suite,
        'build_multi_metric_chart_result': build_multi_metric_chart_result,
        'raise_no_data_error': raise_no_data_error,
        'request_retry': request_retry,
        'save_empty_analysis_result': save_empty_analysis_result,
        'save_analysis_result': save_analysis_result,
    }


def _serialize_runtime_user_context(runtime_context: CustomContext | None) -> dict[str, Any]:
    """从 Agent 运行时上下文提取对子进程执行真正有用的最小信息。"""
    if runtime_context is None:
        return {}

    return {
        'token': getattr(runtime_context, 'auth_token', '') or '',
        'database_conn_info': dict(getattr(runtime_context, 'database_conn_info', {}) or {}),
    }


def _prime_worker_runtime_context(runtime_user_context: dict[str, Any] | None) -> None:
    """在子进程里恢复数据库访问所需的最小运行时上下文。"""
    payload = runtime_user_context or {}
    conn_info_payload = payload.get('database_conn_info') or {}
    if conn_info_payload:
        supos_kernel_api._database_conn_info = DatabaseConnInfo(
            host=str(conn_info_payload.get('host') or ''),
            port=str(conn_info_payload.get('port') or ''),
            user=str(conn_info_payload.get('user') or ''),
            password=str(conn_info_payload.get('password') or ''),
            lake_rds_database_name=str(conn_info_payload.get('lake_rds_database_name') or ''),
        )
        supos_kernel_api._database_pool = None
        return

    token = str(payload.get('token') or '').strip()
    if token:
        supos_kernel_api.get_database_conn_info(token)


def _execute_generated_code_worker(
    code: str,
    username: str,
    runtime_user_context: dict[str, Any],
    result_queue: mp.Queue,
) -> None:
    _prime_worker_runtime_context(runtime_user_context)
    namespace = _build_execution_namespace()
    exec_result: Any = None
    with _capture_runtime_io(username) as (stdout_buffer, stderr_buffer):
        try:
            exec(compile(code, GENERATED_CODE_FILENAME, 'exec'), namespace)
            exec_result = namespace.get('result')
            result_file_path = _write_worker_result_payload({
                'status': 'success',
                'stdout_text': _sanitize_tool_output(stdout_buffer.getvalue()),
                'stderr_text': _sanitize_tool_output(stderr_buffer.getvalue()),
                'exec_result': _serialize_exec_result(exec_result),
            })
            result_queue.put({
                'status': 'success',
                'result_file_path': result_file_path,
            })
        except Exception as exc:
            raw_error_message = str(exc)
            full_traceback = traceback.format_exc()
            code_error_context = _build_generated_code_error_context(
                code=code,
                error=exc,
                traceback_text=full_traceback,
            )
            error_message = _format_error_message_with_code_location(
                raw_error_message,
                code_error_context,
            )
            error_type = _classify_execution_error(exc, raw_error_message)
            error_signature = _extract_error_signature(error_type, error_message)
            repair_instructions = _build_repair_instructions(error_type)
            if isinstance(exc, NoDataFoundError):
                no_data_feedback = _build_no_data_feedback(
                    reason=exc.reason,
                    detail_lines=exc.detail_lines,
                )
                error_message = no_data_feedback['error_message']
                error_signature = no_data_feedback['error_signature']
                repair_instructions = no_data_feedback['repair_instructions']
            result_file_path = _write_worker_result_payload({
                'status': 'error',
                'error_message': error_message,
                'error_type': error_type,
                'error_signature': error_signature,
                'code_error_context': code_error_context,
                'repair_instructions': repair_instructions,
                'stdout_text': _sanitize_tool_output(stdout_buffer.getvalue()),
                'stderr_text': _sanitize_tool_output(
                    '\n'.join(item for item in [stderr_buffer.getvalue(), full_traceback] if item)
                ),
            })
            result_queue.put({
                'status': 'error',
                'result_file_path': result_file_path,
            })


@contextmanager
def _capture_runtime_io(username: str):
    """临时接管 stdout/stderr，并绑定当前运行时上下文变量。"""
    import sys as sys_module

    old_stdout = sys_module.stdout
    old_stderr = sys_module.stderr
    sys_module.stdout = StringIO()
    sys_module.stderr = StringIO()
    username_token = CURRENT_USERNAME.set(username or 'anonymous')

    try:
        yield sys_module.stdout, sys_module.stderr
    finally:
        sys_module.stdout = old_stdout
        sys_module.stderr = old_stderr
        CURRENT_USERNAME.reset(username_token)


def _get_stream_emitter():
    """返回一个尽力可用的 LangGraph 自定义流事件发送器。"""
    try:
        writer = get_stream_writer()
    except Exception:
        writer = None

    def emit(event_type: str, **payload):
        if writer:
            writer({'type': event_type, **payload})

    return emit


def _classify_execution_error(error: Exception | None, error_message: str) -> str:
    """把执行失败归类成更容易被模型理解的错误类型。"""
    if isinstance(error, SyntaxError):
        return 'syntax_error'
    if isinstance(error, ModuleNotFoundError):
        return 'missing_dependency'
    if isinstance(error, NoDataFoundError):
        return 'no_data_found'
    if isinstance(error, FileNotFoundError):
        return 'data_source_not_found'
    if isinstance(error, KeyError):
        return 'schema_or_column_mismatch'
    if isinstance(error, NameError):
        return 'undefined_name'
    if isinstance(error, MemoryError):
        return 'resource_exhausted'

    lowered = (error_message or '').lower()
    if 'memoryerror' in lowered or 'cannot allocate memory' in lowered or 'out of memory' in lowered:
        return 'resource_exhausted'
    if _is_chart_contract_error_message(error_message):
        return 'chart_contract_error'
    if 'got an unexpected keyword argument' in lowered:
        return 'library_api_signature_mismatch'
    if 'invalid comparison' in lowered or ('datetime' in lowered and 'timestamp' in lowered):
        return 'data_type_or_time_mismatch'
    if 'not in index' in lowered or 'no such column' in lowered:
        return 'schema_or_column_mismatch'
    if (
        'no_data_found' in lowered
        or '未命中可分析数据' in error_message
        or '筛选结果为空' in error_message
        or '过滤后无数据' in error_message
        or '关联后无数据' in error_message
    ):
        return 'no_data_found'
    if 'save_analysis_result' in lowered or 'analysis_report' in lowered or 'result' in lowered:
        return 'result_contract_error'
    return 'runtime_error'


def _extract_error_signature(error_type: str, error_message: str) -> dict[str, Any]:
    """从错误信息中提取适合下一轮修补使用的错误签名。"""
    signature: dict[str, Any] = {
        'error_type': error_type,
        'message': error_message or '',
    }

    keyword_match = re.search(
        r"(?P<owner>[A-Za-z_][\w.]*)\.__init__\(\) got an unexpected keyword argument '(?P<argument>[^']+)'",
        error_message or '',
    )
    if keyword_match:
        signature.update({
            'signature_type': 'unexpected_keyword_argument',
            'owner': keyword_match.group('owner'),
            'argument': keyword_match.group('argument'),
        })
        return signature

    module_match = re.search(r"No module named '([^']+)'", error_message or '')
    if module_match:
        signature.update({
            'signature_type': 'missing_dependency',
            'module_name': module_match.group(1),
        })
        return signature

    key_match = re.search(r"'([^']+)'", error_message or '')
    if error_type == 'schema_or_column_mismatch' and key_match:
        signature.update({
            'signature_type': 'schema_or_column_mismatch',
            'field_name': key_match.group(1),
        })
        return signature

    if error_type == 'data_type_or_time_mismatch':
        signature['signature_type'] = 'data_type_or_time_mismatch'
        return signature

    if error_type == 'resource_exhausted':
        exitcode_match = re.search(r'exitcode=(-?\d+)', error_message or '')
        signature.update({
            'signature_type': 'resource_exhausted',
            'exitcode': int(exitcode_match.group(1)) if exitcode_match else None,
        })
        return signature

    if error_type == 'no_data_found':
        signature['signature_type'] = 'no_data_found'
        return signature

    if error_type == 'chart_contract_error':
        signature['signature_type'] = 'chart_contract_error'
        return signature

    signature['signature_type'] = 'generic'
    return signature


def _build_turn_failure_memory(turn_id: int, current_execution_id: int) -> dict[str, Any]:
    """汇总同一轮中之前已失败的执行记录，供模型避免重复犯错。"""
    if turn_id <= 0:
        return {
            'previous_failure_messages': [],
            'previous_failure_hints': [],
        }

    from config.database import SessionLocal
    from model import InsightNsExecution

    session = SessionLocal()
    try:
        previous_failures = session.query(InsightNsExecution).filter(
            InsightNsExecution.turn_id == turn_id,
            InsightNsExecution.id != current_execution_id,
            InsightNsExecution.is_deleted == 0,
            InsightNsExecution.execution_status != 'success',
        ).order_by(InsightNsExecution.created_at.asc()).all()

        failure_messages: list[str] = []
        failure_hints: list[str] = []
        seen_hints: set[str] = set()

        for execution in previous_failures:
            error_message = (execution.error_message or '').strip()
            if not error_message:
                continue

            failure_messages.append(error_message)
            error_signature = _extract_error_signature(
                _classify_execution_error(None, error_message),
                error_message,
            )
            signature_type = error_signature.get('signature_type')

            if signature_type == 'unexpected_keyword_argument':
                owner = error_signature.get('owner', '当前库')
                argument = error_signature.get('argument', '')
                hint = f'之前已经确认 {owner} 不支持参数 `{argument}`，不要再次生成这段写法。'
            elif signature_type == 'missing_dependency':
                module_name = error_signature.get('module_name', '')
                hint = f'之前已经确认当前环境缺少模块 `{module_name}`，不要再次 import 它。'
            elif signature_type == 'schema_or_column_mismatch':
                field_name = error_signature.get('field_name', '')
                hint = f'之前已经出现字段或列不匹配问题（{field_name}），请先核对真实列名。'
            elif signature_type == 'data_type_or_time_mismatch':
                hint = '之前已经出现时间或数据类型比较不兼容问题，请先统一类型后再过滤。'
            elif signature_type == 'resource_exhausted':
                hint = '之前已经出现内存或资源耗尽问题，请优先改用 load_local_file_low_memory(...) 分批读取，并在批次内完成过滤、聚合或累计统计。'
            elif signature_type == 'no_data_found':
                hint = '之前已经出现“筛选/关联后无数据”，请优先检查 SQL、时间范围、筛选条件或 JOIN 条件，不要继续输出空图表。'
            else:
                hint = f'不要再次重复已出现过的错误：{error_message[:120]}'

            if hint not in seen_hints:
                seen_hints.add(hint)
                failure_hints.append(hint)

        return {
            'previous_failure_messages': failure_messages[-3:],
            'previous_failure_hints': failure_hints[-3:],
        }
    finally:
        session.close()


def _build_repair_instructions(error_type: str) -> list[str]:
    """根据错误类型生成定向修复建议。"""
    common_instructions = [
        '重新生成完整可执行代码，不要只修改局部片段。',
        '正式分析代码应调用 save_analysis_result(...) 并把返回值赋给 result；探测或结构化反馈代码应调用 request_retry(...) 并把返回值赋给 result。',
        '如果当前修正思路与上一次相同且未解决错误，请更换实现方式后再重试。',
    ]

    specific_instructions = {
        'syntax_error': [
            '先检查引号、括号、缩进和多行字符串是否闭合，再重新生成代码。',
            '构造分析报告时优先使用 report_sections/report_lines 加 "\\n\\n".join(...)，避免手写长字符串拼接。',
        ],
        'missing_dependency': [
            '不要依赖执行环境中未注入或未安装的第三方模块。',
            '优先使用当前工具已提供的内置辅助函数，而不是额外 import 新库。',
        ],
        'data_source_not_found': [
            '检查数据源标识、文件路径、表名或接口地址是否与当前数据源上下文一致。',
            '不要凭空假设新的文件路径或表名。',
            '这类错误表示必要数据源不可访问或标识不匹配，不属于 no_data_found；不要调用 raise_no_data_error(...) 包装文件不存在、表不存在或接口不可访问。',
            '如果已原样使用当前数据源上下文中的标识但仍不可访问，应调用 request_retry(retry_type="data_source_unavailable", ...) 返回结构化反馈，或让本轮失败，不要伪装成空数据分析。',
        ],
        'schema_or_column_mismatch': [
            '字段选择必须以 metadata_schema.properties 中提供的真实字段名为准。',
            '重新检查 DataFrame 列名后再生成过滤、聚合和展示逻辑。',
            '如果过滤后或关联后的结果为空，不要继续修改字段名；应先判断是否为空，并调用 raise_no_data_error(reason=..., detail_lines=[...]) 把信息返回给上层重试。',
        ],
        'undefined_name': [
            '重新检查变量名、函数名和 import 是否一致，避免引用未定义对象。',
            '尽量复用前面已经定义好的中间变量，不要混用多个命名。',
        ],
        'data_type_or_time_mismatch': [
            '先统一时间列或比较字段的数据类型，再进行过滤或比较。',
            '遇到相对日期或时区处理时，优先考虑使用 get_day_range() 并保证两侧都是可比较的带时区时间对象。',
            '如果目标时间窗口没有命中数据，不要输出空图表或直接调用 save_empty_analysis_result() 收口；应调用 raise_no_data_error(reason=..., detail_lines=[...]) 把信息返回给上层重试。',
        ],
        'no_data_found': [
            '请先复核 SQL、时间范围、筛选条件、JOIN 条件或聚合口径是否过严，再生成下一版完整代码。',
            '在数据加载、过滤、关联、聚合完成后，必须先检查 DataFrame 是否为空；只要为空，就调用 raise_no_data_error(reason=..., detail_lines=[...])，不要继续生成空图表或空结果产物。',
            'reason 要直接说明是哪一步没有命中数据，detail_lines 可补充当前时间范围、筛选条件、关联键或中间行数，方便上层感知后继续重试。',
        ],
        'probe_feedback': [
            '当前反馈是数据探测结果，不是最终分析报告；请基于 diagnostics 中的候选值、时间覆盖或字段分布重写正式分析代码。',
            '下一版代码应切换为 execution_intent = "analysis"，并调用 save_analysis_result(...) 返回最终分析结果。',
            '不要再次只输出探测报告；如果候选值仍有歧义，再返回 request_retry(retry_type="probe_feedback", ...)。',
        ],
        'data_source_unavailable': [
            '当前反馈表示必要数据源不可访问或标识无法解析，不是筛选条件未命中数据。',
            '不要调用 raise_no_data_error(...)，也不要继续生成空图表或空报告伪装成功。',
            '只能原样使用当前数据源上下文提供的文件路径、表名或接口地址；如果仍不可访问，应保持 data_source_unavailable 语义并让上层失败收口。',
        ],
        'missing_structured_artifact': [
            '当前代码已经生成分析报告，但没有生成 charts 或 tables 结构化产物。',
            '下一版代码应保留已有统计口径、筛选条件和核心结论，至少补充一个 charts 或 tables。',
            '趋势、对比、TopN 等场景优先生成 charts；单指标或明细汇总可以生成 tables。',
        ],
        'missing_report': [
            '当前结果缺少非空 analysis_report；请先构造完整 Markdown 分析报告，再调用 save_analysis_result(...)。',
            '不要只返回图表或表格，最终结果必须同时包含可读报告和结构化产物。',
        ],
        'incomplete_result': [
            '当前执行结果不完整；请根据 diagnostics 补齐缺失的报告、图表或表格后重新生成完整代码。',
            '不要把中间探测信息包装成最终分析报告。',
        ],
        'contract_error': [
            '当前代码主动返回了契约问题反馈；请根据 message、diagnostics 和 repair_instructions 重新生成完整代码。',
            '如果本轮是探测，请返回 request_retry(...)；如果本轮是正式分析，请返回 save_analysis_result(...)。',
        ],
        'chart_contract_error': [
            '当前失败是图表产物契约错误，不是数据查询失败；不要改成 no_data_found，也不要只输出自然语言报告。',
            '下一版禁止继续使用 matplotlib/base64 图片、手写 chart_spec 或自造 chart_type/chart_kind 结构。',
            '必须优先用 build_chart_result(...)、build_multi_metric_chart_result(...) 或 build_chart_suite(...) 生成 charts；bar/line 需要 category_field 和 value_field，多指标柱图需要 category_field 和 value_fields，pie 需要 category_field 和 value_field，scatter 需要 x_field 和 y_field。',
            'charts 应直接等于 helper 的返回结果，例如 charts = [build_chart_result(...)]、charts = [build_multi_metric_chart_result(...)] 或 charts = build_chart_suite(...)，然后调用 save_analysis_result(analysis_report=..., charts=charts, tables=...)。',
            '如果确实不适合生成图表，也必须用 tables 返回结构化汇总，不能让 charts 和 tables 同时为空。',
        ],
        'result_contract_error': [
            '如果错误与 charts/chart_spec/chart_document 有关，下一版不要手写图表 JSON，直接改用 build_chart_result(...)、build_multi_metric_chart_result(...) 或 build_chart_suite(...)。',
            '最终必须产出完整 analysis_report，并调用 save_analysis_result(analysis_report=..., charts=[...], tables=[...])。',
            '不要遗漏 result 变量，也不要返回空的 analysis_report 或 charts/tables 同时为空的结构化结果。',
        ],
        'library_api_signature_mismatch': [
            '如果报错来自 pyecharts/matplotlib 或图表参数，优先改用 build_chart_result(...)、build_multi_metric_chart_result(...) 或 build_chart_suite(...)，不要继续试底层图表 API。',
            '当前错误属于库函数签名不兼容；请只修正报错位置的非法参数，不要重写与该错误无关的数据处理和图表逻辑。',
            '如果某个 opts 或图表参数不被当前版本支持，请删除该非法参数，或改成更基础、更稳定的默认写法。',
            '优先保留图表结构、数据处理和报告逻辑，只对报错的 API 调用做最小修改。',
        ],
        'resource_exhausted': [
            '当前失败更像是大文件加载或内存耗尽，不要再一次性把整份本地文件完整加载到 pandas 内存中。',
            '如果是本地文件，请改用 load_local_file_low_memory(...) 分批读取；可以先取第一个小批次确认列名、类型、时间列和候选取值，再继续正式分析。',
            '优先只读取必要列，并在每个批次内完成过滤、聚合、TopN 累计或有限明细截取，不要把所有批次重新拼成一个超大 DataFrame。',
            'CSV 和 Excel 都优先走 load_local_file_low_memory(...)；其中 .xlsx/.xlsm 支持流式批量读取，.xls 只能 best-effort 低内存处理。',
            '如果用户要的是聚合、趋势、分布、TopN 或有限明细，就在批次循环中累计最终结果，只保留用于图表和报告生成的小结果。',
            '如果旧版 .xls 文件在当前环境仍无法安全低内存处理，应明确返回受限说明或建议转换为 .xlsx/CSV，不要继续反复 OOM 重试。',
        ],
        'runtime_error': [
            '优先查看 error_message 中的“错误定位”和 code_error_context.line_no/code_error_context.context_lines，直接修正生成代码中对应行附近的逻辑。',
            '如果 code_error_context 中给出了 traceback_excerpt，请结合最后一帧生成代码位置判断根因，不要只根据异常类型泛泛重写。',
            '如果错误发生在数据处理阶段，请先验证中间结果再进入图表和报告生成。',
        ],
        'execution_timeout': [
            '当前代码执行超时，说明存在无界循环、超大范围数据处理或阻塞式等待。',
            '请先缩小数据范围、减少无必要的明细拼接与循环处理，并确保所有外部请求都具备明确超时。',
            '如果当前分析使用的是表数据源，优先把时间过滤、字段裁剪、JOIN、GROUP BY、LIMIT 下推到 SQL 层完成，不要继续把大结果集加载到 pandas 后再 merge 或聚合。',
            '如果跨表 JOIN 或明细关联已经连续超时，请先退化成单表分析：优先完成主表趋势、分布和有限明细，再在分析报告中明确说明关联部分受查询成本限制暂未展开。',
            '如果需要展示明细，请先限制结果行数，再生成图表和分析报告。',
        ],
    }
    return [*specific_instructions.get(error_type, specific_instructions['runtime_error']), *common_instructions]


def _tool_error_message(
    *,
    tool_call_id: str,
    error_type: str,
    error_message: str,
    repair_instructions: list[str],
    error_signature: dict[str, Any] | None = None,
    code_error_context: dict[str, Any] | None = None,
    feedback_kind: str = '',
    retry_reason_code: str = '',
    diagnostics: dict[str, Any] | None = None,
    analysis_report: str = '',
    current_failed_code: str = '',
    previous_failure_messages: list[str] | None = None,
    previous_failure_hints: list[str] | None = None,
    stdout_text: str = '',
    stderr_text: str = '',
) -> ToolMessage:
    """构造统一的结构化工具错误反馈，供 Agent 循环继续修正。"""
    clipped_current_code = _clip_retry_context_text(current_failed_code, TOOL_ERROR_CURRENT_CODE_MAX_CHARS)
    clipped_error_message = _clip_retry_context_text(error_message, TOOL_ERROR_TEXT_MAX_CHARS)
    clipped_error_signature = dict(error_signature or {})
    if 'message' in clipped_error_signature:
        clipped_error_signature['message'] = _clip_retry_context_text(
            clipped_error_signature.get('message'),
            TOOL_ERROR_TEXT_MAX_CHARS,
        )
    clipped_code_error_context = dict(code_error_context or {})
    if clipped_code_error_context.get('traceback_excerpt'):
        clipped_code_error_context['traceback_excerpt'] = _clip_retry_context_text(
            clipped_code_error_context.get('traceback_excerpt'),
            TOOL_ERROR_CODE_CONTEXT_MAX_CHARS,
        )

    payload = {
        'tool': 'execute_python',
        'status': 'failed',
        'error_type': error_type,
        'error_message': clipped_error_message,
        'error_message_truncated': clipped_error_message != (error_message or ''),
        'repair_instructions': _clip_retry_context_list(
            repair_instructions,
            max_items=8,
            max_chars=TOOL_ERROR_INSTRUCTION_MAX_CHARS,
        ),
        'error_signature': clipped_error_signature,
        'code_error_context': clipped_code_error_context,
        'minimal_patch_required': True,
        'current_failed_code': clipped_current_code,
        'current_failed_code_truncated': clipped_current_code != (current_failed_code or ''),
        'previous_failure_messages': _clip_retry_context_list(
            previous_failure_messages,
            max_items=3,
            max_chars=TOOL_ERROR_PREVIOUS_ITEM_MAX_CHARS,
        ),
        'previous_failure_hints': _clip_retry_context_list(
            previous_failure_hints,
            max_items=3,
            max_chars=TOOL_ERROR_PREVIOUS_ITEM_MAX_CHARS,
        ),
        'previous_failure_count': len(previous_failure_messages or []),
    }
    if feedback_kind:
        payload['feedback_kind'] = feedback_kind
    if retry_reason_code:
        payload['retry_reason_code'] = retry_reason_code
    if diagnostics:
        payload['diagnostics'] = diagnostics
    if analysis_report:
        payload['analysis_report'] = _clip_retry_context_text(analysis_report, TOOL_ERROR_TEXT_MAX_CHARS)
    if stdout_text:
        payload['stdout_text'] = _clip_retry_context_text(stdout_text, TOOL_ERROR_STDIO_MAX_CHARS)
    if stderr_text:
        payload['stderr_text'] = _clip_retry_context_text(stderr_text, TOOL_ERROR_STDIO_MAX_CHARS)

    return ToolMessage(
        content=json.dumps(payload, ensure_ascii=False),
        tool_call_id=tool_call_id,
    )


def _tool_success_message(
    *,
    tool_call_id: str,
    analysis_report: str,
    charts: list[dict[str, Any]] | None = None,
    tables: list[dict[str, Any]] | None = None,
) -> ToolMessage:
    """
    构造成功的工具结果反馈。

    这里不再把完整 charts/tables 配置原样塞回模型上下文，
    避免结构化图表配置把后续一轮模型请求撑爆。
    完整结果已经持久化到 execution.result_payload_json 中，
    后续由 invoker 基于最新成功执行记录回填真正结果。
    """
    payload = {
        'tool': 'execute_python',
        'status': 'success',
        'analysis_report': (analysis_report or '').strip(),
        'chart_count': len(charts or []),
        'table_count': len(tables or []),
        'chart_titles': [
            str(item.get('title') or '').strip()
            for item in (charts or [])
            if isinstance(item, dict) and str(item.get('title') or '').strip()
        ][:5],
        'table_titles': [
            str(item.get('title') or '').strip()
            for item in (tables or [])
            if isinstance(item, dict) and str(item.get('title') or '').strip()
        ][:5],
    }
    return ToolMessage(
        content=json.dumps(payload, ensure_ascii=False),
        tool_call_id=tool_call_id,
    )


@tool(description='执行 Python 代码工具', args_schema=ExePythonCodeInput)
def execute_python(
    runtime: ToolRuntime[CustomContext],
    code: str,
    title: str = '',
    description: str = "",
) -> Optional[ToolMessage]:
    """
    执行模型生成的 Python 分析代码。

    需要注意：
    - 这里保持当前对外工具契约不变。
    - 执行历史仍会内部持久化到 `insight_ns_execution`。
    """
    emit = _get_stream_emitter()
    start_time = time.time()
    execution_id = _create_execution_record(
        conversation_id=int(getattr(runtime.context, 'conversation_id', 0) or 0),
        turn_id=int(getattr(runtime.context, 'turn_id', 0) or 0),
        tool_call_id=getattr(runtime, 'tool_call_id', '') or '',
        code=code,
        title=title,
        description=description,
    )
    execution_log_context = {
        'username': getattr(runtime.context, 'username', 'anonymous'),
        'namespace_id': int(getattr(runtime.context, 'namespace_id', 0) or 0),
        'conversation_id': int(getattr(runtime.context, 'conversation_id', 0) or 0),
        'turn_id': int(getattr(runtime.context, 'turn_id', 0) or 0),
        'execution_id': execution_id,
    }

    task_title = title or '数据分析任务'
    emit(
        'status',
        stage='tool_start',
        level='info',
        tool='execute_python',
        message=f"正在执行分析代码：{task_title}",
    )
    with logger.context(**execution_log_context):
        logger.info(f"开始执行：{task_title}")
        logger.info(f"代码：\n```python\n{code}\n```")
        logger.info("代码已提交到本地执行器。")
    emit(
        'status',
        stage='tool_running',
        level='info',
        tool='execute_python',
        message='分析代码已提交到本地执行器。',
    )

    from config.config import Config

    timeout_seconds = max(int(getattr(Config, 'PYTHON_EXEC_TIMEOUT_SECONDS', 90) or 90), 1)
    result_queue: mp.Queue = mp.Queue()
    runtime_user_context = _serialize_runtime_user_context(runtime.context)
    process = mp.Process(
        target=_execute_generated_code_worker,
        args=(code, getattr(runtime.context, 'username', 'anonymous'), runtime_user_context, result_queue),
        daemon=True,
    )

    worker_result: dict[str, Any] | None = None
    stdout_text = ''
    stderr_text = ''
    exec_result: Any = None
    process.start()
    deadline = time.time() + timeout_seconds

    while time.time() < deadline:
        remaining_seconds = max(deadline - time.time(), 0)
        try:
            worker_result = result_queue.get(timeout=min(0.5, remaining_seconds))
            break
        except Empty:
            if not process.is_alive():
                break

    if worker_result is None and process.is_alive():
        process.terminate()
        process.join(5)
        error_message = f'代码执行超时，已超过 {timeout_seconds} 秒'
        error_type = 'execution_timeout'
        repair_instructions = _build_repair_instructions(error_type)
        error_signature = _extract_error_signature(error_type, error_message)
        failure_memory = _build_turn_failure_memory(
            turn_id=int(getattr(runtime.context, 'turn_id', 0) or 0),
            current_execution_id=execution_id,
        )
        _update_execution_record(
            execution_id,
            execution_status='failed',
            stdout_text='',
            stderr_text='',
            execution_seconds=int((time.time() - start_time) * 1000),
            error_message=error_message,
        )
        with logger.context(**execution_log_context):
            logger.error(error_message)
        emit(
            'status',
            stage='tool_error',
            level='error',
            tool='execute_python',
            message=error_message,
        )
        return _tool_error_message(
            tool_call_id=runtime.tool_call_id,
            error_type=error_type,
            error_message=error_message,
            repair_instructions=repair_instructions,
            error_signature=error_signature,
            current_failed_code=code,
            previous_failure_messages=failure_memory.get('previous_failure_messages', []),
            previous_failure_hints=failure_memory.get('previous_failure_hints', []),
        )

    process.join(5)
    if process.is_alive():
        process.terminate()
        process.join(5)
    process_exitcode = process.exitcode
    process.close()
    result_queue.close()
    result_queue.join_thread()

    if worker_result is None:
        error_type, error_message = _classify_process_exit(process_exitcode)
        repair_instructions = _build_repair_instructions(error_type)
        error_signature = _extract_error_signature(error_type, error_message)
        failure_memory = _build_turn_failure_memory(
            turn_id=int(getattr(runtime.context, 'turn_id', 0) or 0),
            current_execution_id=execution_id,
        )
        _update_execution_record(
            execution_id,
            execution_status='failed',
            stdout_text='',
            stderr_text='',
            execution_seconds=int((time.time() - start_time) * 1000),
            error_message=error_message,
        )
        with logger.context(**execution_log_context):
            logger.error(error_message)
        emit(
            'status',
            stage='tool_error',
            level='error',
            tool='execute_python',
            message=error_message,
        )
        return _tool_error_message(
            tool_call_id=runtime.tool_call_id,
            error_type=error_type,
            error_message=error_message,
            repair_instructions=repair_instructions,
            error_signature=error_signature,
            current_failed_code=code,
            previous_failure_messages=failure_memory.get('previous_failure_messages', []),
            previous_failure_hints=failure_memory.get('previous_failure_hints', []),
        )

    result_file_path = str(worker_result.get('result_file_path') or '').strip()
    try:
        worker_payload = _read_worker_result_payload(result_file_path) if result_file_path else worker_result
    except Exception as exc:
        error_message = f'代码执行结果回读失败: {exc}'
        error_type = 'runtime_error'
        repair_instructions = _build_repair_instructions(error_type)
        error_signature = _extract_error_signature(error_type, error_message)
        failure_memory = _build_turn_failure_memory(
            turn_id=int(getattr(runtime.context, 'turn_id', 0) or 0),
            current_execution_id=execution_id,
        )
        _update_execution_record(
            execution_id,
            execution_status='failed',
            stdout_text='',
            stderr_text='',
            execution_seconds=int((time.time() - start_time) * 1000),
            error_message=error_message,
        )
        with logger.context(**execution_log_context):
            logger.error(error_message)
        emit(
            'status',
            stage='tool_error',
            level='error',
            tool='execute_python',
            message=error_message,
        )
        return _tool_error_message(
            tool_call_id=runtime.tool_call_id,
            error_type=error_type,
            error_message=error_message,
            repair_instructions=repair_instructions,
            error_signature=error_signature,
            current_failed_code=code,
            previous_failure_messages=failure_memory.get('previous_failure_messages', []),
            previous_failure_hints=failure_memory.get('previous_failure_hints', []),
        )
    finally:
        _remove_worker_result_payload(result_file_path)

    stdout_text = _sanitize_tool_output(worker_payload.get('stdout_text', ''))
    stderr_text = _sanitize_tool_output(worker_payload.get('stderr_text', ''))
    if worker_payload.get('status') == 'error':
        error_message = worker_result.get('error_message', '代码执行失败')
        error_type = worker_result.get('error_type', 'runtime_error')
        repair_instructions = worker_result.get('repair_instructions') or _build_repair_instructions(error_type)
        error_signature = worker_result.get('error_signature') or _extract_error_signature(error_type, error_message)
        error_message = worker_payload.get('error_message', error_message)
        error_type = worker_payload.get('error_type', error_type)
        repair_instructions = worker_payload.get('repair_instructions') or _build_repair_instructions(error_type)
        error_signature = worker_payload.get('error_signature') or _extract_error_signature(error_type, error_message)
        code_error_context = worker_payload.get('code_error_context') if isinstance(worker_payload.get('code_error_context'), dict) else {}
        failure_memory = _build_turn_failure_memory(
            turn_id=int(getattr(runtime.context, 'turn_id', 0) or 0),
            current_execution_id=execution_id,
        )
        _update_execution_record(
            execution_id,
            execution_status='failed',
            result_payload_json=json.dumps({
                'status': 'failed',
                'error_type': error_type,
                'error_message': error_message,
                'error_signature': error_signature,
                'code_error_context': code_error_context,
                'repair_instructions': repair_instructions,
            }, ensure_ascii=False),
            stdout_text=stdout_text,
            stderr_text=stderr_text,
            execution_seconds=int((time.time() - start_time) * 1000),
            error_message=error_message,
        )
        with logger.context(**execution_log_context):
            logger.error(f"代码执行失败（{error_type}）：{error_message}")
        emit(
            'status',
            stage='tool_error',
            level='error',
            tool='execute_python',
            message=f'代码执行失败（{error_type}）：{error_message}',
        )
        return _tool_error_message(
            tool_call_id=runtime.tool_call_id,
            error_type=error_type,
            error_message=error_message,
            repair_instructions=repair_instructions,
            error_signature=error_signature,
            code_error_context=code_error_context,
            current_failed_code=code,
            previous_failure_messages=failure_memory.get('previous_failure_messages', []),
            previous_failure_hints=failure_memory.get('previous_failure_hints', []),
            stdout_text=stdout_text,
            stderr_text=stderr_text,
        )

    exec_result = _deserialize_exec_result(worker_result.get('exec_result') or {})
    exec_result = _deserialize_exec_result(worker_payload.get('exec_result') or {})

    execution_seconds = int((time.time() - start_time) * 1000)
    with logger.context(**execution_log_context):
        logger.info(f"执行完成，耗时 {execution_seconds / 1000:.2f} 秒")
    emit(
        'status',
        stage='tool_finished',
        level='success',
        tool='execute_python',
        message=f'代码执行完成，耗时 {execution_seconds / 1000:.2f} 秒',
    )

    if stdout_text:
        # stdout/stderr 既用于调试，也用于后续回看执行历史。
        with logger.context(**execution_log_context):
            logger.info(stdout_text)
        emit(
            'tool_log',
            stage='tool_output',
            level='info',
            tool='execute_python',
            message=stdout_text[:1000],
        )
    if stderr_text:
        with logger.context(**execution_log_context):
            logger.info(f"标准错误：\n{stderr_text}")
        emit(
            'tool_log',
            stage='tool_output',
            level='warning',
            tool='execute_python',
            message=stderr_text[:1000],
        )

    if isinstance(exec_result, RetryResult):
        retry_error_signature = _build_retry_error_signature(exec_result)
        retry_repair_instructions = exec_result.repair_instructions or _build_repair_instructions(exec_result.retry_type)
        failure_memory = _build_turn_failure_memory(
            turn_id=int(getattr(runtime.context, 'turn_id', 0) or 0),
            current_execution_id=execution_id,
        )
        _update_execution_record(
            execution_id,
            execution_status='failed',
            analysis_report=exec_result.analysis_report,
            result_payload_json=exec_result.model_dump_json(),
            stdout_text=stdout_text,
            stderr_text=stderr_text,
            execution_seconds=execution_seconds,
            error_message=exec_result.message,
        )
        emit(
            'status',
            stage='tool_retry',
            level='warning',
            tool='execute_python',
            message=exec_result.message,
        )
        return _tool_error_message(
            tool_call_id=runtime.tool_call_id,
            error_type=exec_result.retry_type,
            error_message=exec_result.message,
            repair_instructions=retry_repair_instructions,
            error_signature=retry_error_signature,
            feedback_kind='retryable_feedback',
            retry_reason_code=exec_result.retry_type,
            diagnostics=exec_result.diagnostics,
            analysis_report=exec_result.analysis_report,
            current_failed_code=code,
            previous_failure_messages=failure_memory.get('previous_failure_messages', []),
            previous_failure_hints=failure_memory.get('previous_failure_hints', []),
            stdout_text=stdout_text,
            stderr_text=stderr_text,
        )

    if isinstance(exec_result, StructuredResult):
        contract_retry = _validate_structured_result_contract(exec_result)
        if contract_retry is not None:
            failure_memory = _build_turn_failure_memory(
                turn_id=int(getattr(runtime.context, 'turn_id', 0) or 0),
                current_execution_id=execution_id,
            )
            _update_execution_record(
                execution_id,
                execution_status='failed',
                analysis_report=contract_retry.analysis_report or exec_result.analysis_report,
                result_payload_json=contract_retry.model_dump_json(),
                stdout_text=stdout_text,
                stderr_text=stderr_text,
                execution_seconds=execution_seconds,
                error_message=contract_retry.message,
            )
            emit(
                'status',
                stage='tool_retry',
                level='warning',
                tool='execute_python',
                message=contract_retry.message,
            )
            return _tool_error_message(
                tool_call_id=runtime.tool_call_id,
                error_type=contract_retry.retry_type,
                error_message=contract_retry.message,
                repair_instructions=contract_retry.repair_instructions or _build_repair_instructions(contract_retry.retry_type),
                error_signature=_build_retry_error_signature(contract_retry),
                feedback_kind='retryable_feedback',
                retry_reason_code=contract_retry.retry_type,
                diagnostics=contract_retry.diagnostics,
                analysis_report=contract_retry.analysis_report,
                current_failed_code=code,
                previous_failure_messages=failure_memory.get('previous_failure_messages', []),
                previous_failure_hints=failure_memory.get('previous_failure_hints', []),
                stdout_text=stdout_text,
                stderr_text=stderr_text,
            )

        # 对外继续保持 ToolMessage JSON 结构，对内持久化完整结构化执行结果。
        _update_execution_record(
            execution_id,
            execution_status='success',
            analysis_report=exec_result.analysis_report,
            result_payload_json=exec_result.model_dump_json(),
            stdout_text=stdout_text,
            stderr_text=stderr_text,
            execution_seconds=execution_seconds,
        )
        emit(
            'status',
            stage='tool_result',
            level='success',
            tool='execute_python',
            message='分析结果已生成，正在整理最终报告。',
        )
        return _tool_success_message(
            tool_call_id=runtime.tool_call_id,
            analysis_report=exec_result.analysis_report,
            charts=exec_result.charts,
            tables=exec_result.tables,
        )

    _update_execution_record(
        execution_id,
        execution_status='invalid',
        stdout_text=stdout_text,
        stderr_text=stderr_text,
        execution_seconds=execution_seconds,
        error_message='生成的代码未按模板产出 result',
    )
    # 如果生成代码没有产出 `result`，这里会要求模型重试，
    # 而不是静默接受一次不完整的执行。
    emit(
        'status',
        stage='tool_retry',
        level='warning',
        tool='execute_python',
        message='生成的代码未按模板返回 result，正在请求模型修正。',
    )
    return _tool_error_message(
        tool_call_id=runtime.tool_call_id,
        error_type='result_contract_error',
        error_message='生成的代码未按模板产出 result',
        repair_instructions=_build_repair_instructions('result_contract_error'),
        error_signature=_extract_error_signature('result_contract_error', '生成的代码未按模板产出 result'),
        current_failed_code=code,
        previous_failure_messages=_build_turn_failure_memory(
            turn_id=int(getattr(runtime.context, 'turn_id', 0) or 0),
            current_execution_id=execution_id,
        ).get('previous_failure_messages', []),
        previous_failure_hints=_build_turn_failure_memory(
            turn_id=int(getattr(runtime.context, 'turn_id', 0) or 0),
            current_execution_id=execution_id,
        ).get('previous_failure_hints', []),
        stdout_text=stdout_text,
        stderr_text=stderr_text,
    )
